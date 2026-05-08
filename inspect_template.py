from openpyxl import load_workbook

wb = load_workbook(
    "template/template_cv.xlsx"
)

ws = wb["NAMA"]

print("\n=== ALL USED CELLS ===\n")

for row in ws.iter_rows():

    for cell in row:

        if cell.value is not None:

            print(
                f"{cell.coordinate}"
                f" | {cell.value}"
            )

print("\n=== MERGED CELLS ===\n")

for merged in ws.merged_cells.ranges:

    print(merged)
