from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.cell.cell import MergedCell

from PIL import Image as PILImage

import os


TEMPLATE_PATH = "template/CV_ASLI.xlsx"


FIELD_MAP = {

    "nomor": "C2",
    "tanggal_cv": "P2",
    "furigana": "F5",
    "nama": "F6",
    "alamat": "F7",
    "tanggal_lahir": "F8",
    "usia": "K8",
    "gender": "N8",
    "phone": "F9",
    "agama": "K9",
    "tinggi": "N9",
    "gol_darah": "F10",
    "status": "I10",
    "anak": "L10",
    "berat": "N10",
    "paspor": "R10",

    "pernah_jepang": "M25",
    "negara_lain": "M26",

    "kontak_nama": "G36",
    "kontak_hubungan": "K36",
    "kontak_nomor": "N36",

    "alasan": "F38",
    "hobi": "F40",
    "keahlian": "M40",
    "kelebihan": "F41",
    "kekurangan": "M41",

    "bisa_kendara_1": "F42",
    "bisa_kendara_2": "F43",
    "bisa_kendara_3": "F44",

    # MERGED CELL
    "sim_1": "M42",
    "sim_2": "M43",
    "sim_3": "M44",

    "jp_duration": "I25",
    "jp_unit": "J25",

    "bahasa": "F26",
    "level_bahasa": "I26",
    "bahasa_unit": "J26",
}


PENDIDIKAN_ROWS = {
    0: 12,
    1: 13,
    2: 14,
    3: 15,
}


PENGALAMAN_ROWS = {
    0: 17,
    1: 19,
    2: 21,
    3: 23,
}


KELUARGA_ROWS = {
    0: 29,
    1: 30,
    2: 31,
    3: 32,
    4: 33,
    5: 34,
    6: 35,
}


# =================================================
# SAFE VALUE
# =================================================

def safe_value(value):

    if value is None:
        return "-"

    value = str(value).strip()

    if value == "":
        return "-"

    return value


# =================================================
# SAFE YEAR
# =================================================

def safe_year(value):

    if value:
        return f"{value}年"

    return "-"


# =================================================
# SAFE MONTH
# =================================================

def safe_month(value):

    if value:
        return f"{value}月"

    return "-"


# =================================================
# WRITE CELL SAFE
# =================================================

def write_cell(ws, cell, value):

    target_cell = ws[cell]

    if isinstance(target_cell, MergedCell):
        return

    target_cell.value = safe_value(value)


# =================================================
# GENERATE CV
# =================================================

def generate_cv(data):

    os.makedirs("output", exist_ok=True)

    wb = load_workbook(TEMPLATE_PATH)

    ws = wb["NAMA"]

    # =================================================
    # FIELD
    # =================================================

    for key, cell in FIELD_MAP.items():

        write_cell(
            ws,
            cell,
            data.get(key, "-")
        )

    # =================================================
    # FOTO
    # =================================================

    foto = data.get("foto")

    if foto:

        temp_path = "output/temp_photo.png"

        final_path = "output/final_photo.png"

        with open(temp_path, "wb") as f:
            f.write(foto.read())

        pil_img = PILImage.open(temp_path)

        if pil_img.mode != "RGB":
            pil_img = pil_img.convert("RGB")

        FRAME_WIDTH = 145
        FRAME_HEIGHT = 200

        img_width, img_height = pil_img.size

        img_ratio = img_width / img_height

        frame_ratio = FRAME_WIDTH / FRAME_HEIGHT

        # =============================================
        # CROP
        # =============================================

        if img_ratio > frame_ratio:

            new_width = int(
                img_height * frame_ratio
            )

            left = (
                img_width - new_width
            ) // 2

            pil_img = pil_img.crop(
                (
                    left,
                    0,
                    left + new_width,
                    img_height
                )
            )

        else:

            new_height = int(
                img_width / frame_ratio
            )

            top = (
                img_height - new_height
            ) // 2

            pil_img = pil_img.crop(
                (
                    0,
                    top,
                    img_width,
                    top + new_height
                )
            )

        # =============================================
        # RESIZE
        # =============================================

        pil_img = pil_img.resize(
            (
                FRAME_WIDTH,
                FRAME_HEIGHT
            )
        )

        pil_img.save(final_path)

        excel_img = Image(final_path)

        excel_img.width = FRAME_WIDTH
        excel_img.height = FRAME_HEIGHT

        ws.add_image(excel_img, "P5")

    # =================================================
    # PENDIDIKAN
    # =================================================

    for index, item in enumerate(
        data.get("pendidikan", [])
    ):

        if index not in PENDIDIKAN_ROWS:
            continue

        row = PENDIDIKAN_ROWS[index]

        write_cell(
            ws,
            f"F{row}",
            safe_year(
                item.get("tahun_mulai")
            )
        )

        write_cell(
            ws,
            f"G{row}",
            safe_month(
                item.get("bulan_mulai")
            )
        )

        write_cell(
            ws,
            f"I{row}",
            safe_year(
                item.get("tahun_selesai")
            )
        )

        write_cell(
            ws,
            f"J{row}",
            safe_month(
                item.get("bulan_selesai")
            )
        )

        level = item.get("level")

        if level == "SD":

            level_text = "小学生\nSD"

        elif level == "SMP":

            level_text = "中学生\nSMP"

        else:

            level_text = item.get(
                "level_jp",
                "-"
            )

        write_cell(
            ws,
            f"K{row}",
            level_text
        )

        write_cell(
            ws,
            f"L{row}",
            item.get("sekolah")
        )

        if row >= 14:

            write_cell(
                ws,
                f"Q{row}",
                item.get("jurusan")
            )

    # =================================================
    # PENGALAMAN
    # =================================================

    for index, item in enumerate(
        data.get("pengalaman", [])
    ):

        if index not in PENGALAMAN_ROWS:
            continue

        row = PENGALAMAN_ROWS[index]

        # =============================================
        # MULAI
        # =============================================

        write_cell(
            ws,
            f"F{row}",
            safe_year(item.get("tm"))
        )

        write_cell(
            ws,
            f"G{row}",
            safe_month(item.get("bm"))
        )

        # =============================================
        # SELESAI
        # =============================================

        write_cell(
            ws,
            f"F{row+1}",
            safe_year(item.get("ts"))
        )

        write_cell(
            ws,
            f"G{row+1}",
            safe_month(item.get("bs"))
        )

        write_cell(
            ws,
            f"H{row}",
            item.get("perusahaan")
        )

        write_cell(
            ws,
            f"L{row}",
            item.get("upah")
        )

        write_cell(
            ws,
            f"N{row}",
            item.get("pekerjaan")
        )

        write_cell(
            ws,
            f"Q{row}",
            item.get("posisi")
        )

    # =================================================
    # RIWAYAT JEPANG
    # =================================================

    if data.get("japan_tm"):

        japan_text = (

            f'{safe_value(data.get("japan_tm"))}年 '

            f'{safe_value(data.get("japan_bm"))}月 ～ '

            f'{safe_value(data.get("japan_ts"))}年 '

            f'{safe_value(data.get("japan_bs"))}月'
        )

        write_cell(
            ws,
            "N25",
            japan_text
        )

    # =================================================
    # NEGARA LAIN
    # =================================================

    negara_nama = data.get(
        "negara_nama"
    )

    if negara_nama:

        write_cell(
            ws,
            "N26",
            f'({negara_nama} 国)'
        )

    # =================================================
    # KELUARGA
    # =================================================

    for index, item in enumerate(
        data.get("keluarga", [])
    ):

        if index not in KELUARGA_ROWS:
            continue

        row = KELUARGA_ROWS[index]

        write_cell(
            ws,
            f"B{row}",
            item.get("hubungan")
        )

        write_cell(
            ws,
            f"D{row}",
            item.get("nama")
        )

        write_cell(
            ws,
            f"I{row}",
            item.get("usia")
        )

        write_cell(
            ws,
            f"K{row}",
            item.get("pekerjaan")
        )

    # =================================================
    # SAVE
    # =================================================

    output_path = "output/cv_jepang.xlsx"

    wb.save(output_path)

    return output_path